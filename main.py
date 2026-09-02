import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import sys
import re
import json
from datetime import date
import openpyxl
import pandas as pd

# ── 매핑 데이터 저장 경로 ──────────────────────────────────────
def get_mapping_path():
    # PyInstaller exe 실행 시 → exe 파일 위치 기준
    # 일반 py 실행 시 → 스크립트 위치 기준
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, 'mapping.json')

def load_mapping_from_file():
    path = get_mapping_path()
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_mapping_to_file(mapping):
    with open(get_mapping_path(), 'w', encoding='utf-8') as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)

# ── 기본 내장 매핑 (최초 실행 시 사용) ───────────────────────
BUILTIN_MAPPING = {
  "11581018580": {
    "cafe24_id": 252,
    "ss_name": "책상용 칸막이 독서실 사무실 하단수납형 공부 집중 종이 파티션(30%할인)",
    "cafe24_name": "독서실 칸막이 하단수납형 (30%할인)"
  },
  "644533425": {
    "cafe24_id": 99,
    "ss_name": "책상용 칸막이 독서실 사무실 하단수납형 공부 집중 종이 파티션(30%할인)",
    "cafe24_name": "독서실 칸막이 하단수납형 (30%할인)"
  },
  "11580985670": {
    "cafe24_id": 251,
    "ss_name": "미니수납거치대 아이패트 책상수납 정리 소품 페이퍼팝 종이가구 (30%할인)",
    "cafe24_name": "아이패드용 미니 수납함 쑥토리지 (30%할인)"
  },
  "725779737": {
    "cafe24_id": 102,
    "ss_name": "미니수납거치대 아이패트 책상수납 정리 소품 페이퍼팝 종이가구 (30%할인)",
    "cafe24_name": "아이패드용 미니 수납함 쑥토리지 (30%할인)"
  },
  "11553018856": {
    "cafe24_id": 254,
    "ss_name": "공간박스 사과상자 큐브 수납 종이로만든 보관함",
    "cafe24_name": "종이수납함 사과상자"
  },
  "9566079487": {
    "cafe24_id": 105,
    "ss_name": "공간박스 사과상자 큐브 수납 종이로만든 보관함",
    "cafe24_name": "종이수납함 사과상자"
  },
  "11465480794": {
    "cafe24_id": 244,
    "ss_name": "[리페이퍼]  종이 자투리 묶음 가격표, 명찰, 분류표  DIY 교구 창의 (20개입, 30개입)",
    "cafe24_name": "[리페이퍼] 종이 자투리 묶음 (20개입, 30개입)"
  },
  "9958827162": {
    "cafe24_id": 212,
    "ss_name": "[리페이퍼]  종이 자투리 묶음 가격표, 명찰, 분류표  DIY 교구 창의 (20개입, 30개입)",
    "cafe24_name": "[리페이퍼] 종이 자투리 묶음 (20개입, 30개입)"
  },
  "11437818605": {
    "cafe24_id": 247,
    "ss_name": "주문 제작 길안내 표지판 방향 화살표 표시",
    "cafe24_name": "길안내 표지판"
  },
  "644559453": {
    "cafe24_id": 100,
    "ss_name": "주문 제작 길안내 표지판 방향 화살표 표시",
    "cafe24_name": "길안내 표지판"
  },
  "11437705895": {
    "cafe24_id": 246,
    "ss_name": "주문 제작 허니콤보드 배너 전시 행사 팝업 종이 페이퍼팝",
    "cafe24_name": "[주문제작] 허니콤 배너"
  },
  "324326179": {
    "cafe24_id": 115,
    "ss_name": "주문 제작 허니콤보드 배너 전시 행사 팝업 종이 페이퍼팝",
    "cafe24_name": "[주문제작] 허니콤 배너"
  },
  "11437665646": {
    "cafe24_id": 249,
    "ss_name": "종이책장 전면 책꽂이 화이트 1단 2개 묶음 가로34 간이 책꽂이 장난감 사무실 수납 정리",
    "cafe24_name": "[가로34] 화이트 종이책장 ㅁㅁㅂㅂ"
  },
  "11436916855": {
    "cafe24_id": 248,
    "ss_name": "주문 제작 종이 POP 전시 행사 팝업 허니콤보드 종이 페이퍼팝",
    "cafe24_name": "종이 POP"
  },
  "12428193823": {
    "cafe24_id": 78,
    "ss_name": "주문 제작 종이 POP 전시 행사 팝업 허니콤보드 종이 페이퍼팝",
    "cafe24_name": "종이 POP"
  },
  "11386188464": {
    "cafe24_id": 240,
    "ss_name": "코너 수납 장식 선반 2단 틈새 모서리장",
    "cafe24_name": "틈새 코너선반 브이"
  },
  "11364020634": {
    "cafe24_id": 242,
    "ss_name": "주문 제작 입간판매대 전시 행사 팝업 허니콤보드 종이 페이퍼팝",
    "cafe24_name": "[주문제작] 입간판 매대"
  },
  "11337490145": {
    "cafe24_id": 238,
    "ss_name": "주문 제작 입간판매대 전시 행사 팝업 허니콤보드 종이 페이퍼팝",
    "cafe24_name": "[주문제작] 입간판 매대"
  },
  "12428193824": {
    "cafe24_id": 78,
    "ss_name": "종이책장 조립 추가부품 세트 (단 연장)",
    "cafe24_name": "종이책장 조립 추가부품 세트 (단 연장)"
  },
  "11301521053": {
    "cafe24_id": 239,
    "ss_name": "책상 위 미니 3단 수납 (2개입) 다용도 공간활용 작은 테이블 화장대 선반",
    "cafe24_name": "종이 미니선반 2개입"
  },
  "2007122794": {
    "cafe24_id": 103,
    "ss_name": "책상 위 미니 3단 수납 (2개입) 다용도 공간활용 작은 테이블 화장대 선반",
    "cafe24_name": "종이 미니선반 2개입"
  },
  "11243242366": {
    "cafe24_id": 236,
    "ss_name": "접이식 스툴 2개입 휴대용 야외 행사 이벤트 미니보조 의자",
    "cafe24_name": "종이 폴딩 스툴 2개입"
  },
  "4658429166": {
    "cafe24_id": 79,
    "ss_name": "접이식 스툴 2개입 휴대용 야외 행사 이벤트 미니보조 의자",
    "cafe24_name": "종이 폴딩 스툴 2개입"
  },
  "11217201817": {
    "cafe24_id": 235,
    "ss_name": "무타공 인테리어 포켓선반 (2개입) 벽걸이  DIY 벽선반",
    "cafe24_name": "포켓선반 2개입"
  },
  "11178869029": {
    "cafe24_id": 234,
    "ss_name": "발받침대 책상 사무실 2단 다리거치대",
    "cafe24_name": "만족 종이 발받침대"
  },
  "10597303805": {
    "cafe24_id": 223,
    "ss_name": "선반 플랫 80 1단 거실장 오픈 TV 장 장식 진열 수납장",
    "cafe24_name": "플랫 선반 거실장 80"
  },
  "10596349372": {
    "cafe24_id": 225,
    "ss_name": "거실장 TV장 플랫 120 1단 선반 종이가구 오픈장식 진열 수납장",
    "cafe24_name": "플랫 선반 거실장 120"
  },
  "10272872665": {
    "cafe24_id": 220,
    "ss_name": "거실장 TV장 플랫 120 1단 선반 종이가구 오픈장식 진열 수납장",
    "cafe24_name": "플랫 선반 거실장 120"
  },
  "10507452138": {
    "cafe24_id": 222,
    "ss_name": "도서 엽서 진열대 관공서 오피스",
    "cafe24_name": "도서 엽서 진열대"
  },
  "10448496323": {
    "cafe24_id": 211,
    "ss_name": "종이 고양이 임시 화장실 리필용 (속박스3개입)",
    "cafe24_name": "고양이 스크래쳐 교체용"
  },
  "644571006": {
    "cafe24_id": 98,
    "ss_name": "종이 고양이 임시 화장실 리필용 (속박스3개입)",
    "cafe24_name": "고양이 스크래쳐 교체용"
  },
  "10119516050": {
    "cafe24_id": 217,
    "ss_name": "종이 화이트보드 돌림판 룰렛 원판 돌리기",
    "cafe24_name": "종이 화이트 보드 돌림판"
  },
  "10097964371": {
    "cafe24_id": 216,
    "ss_name": "종이 모니터 햇빛 가리개 모니터 가림막 후드 보호 27인치",
    "cafe24_name": "업무집중 모니터가리개"
  },
  "10097857115": {
    "cafe24_id": 215,
    "ss_name": "매거진랙 전면 스탠드 종이 책장 인테리어 수납장 진열 선반 게시대",
    "cafe24_name": "종이 매거진랙 스탠드형 전면책장"
  },
  "12574397828": {
    "cafe24_id": 212,
    "ss_name": "전면형 이동식트레이 종이가구 이동식수납장 간이진열대",
    "cafe24_name": "종이수납장 수수선반 2단 협탁"
  },
  "9681924869": {
    "cafe24_id": 210,
    "ss_name": "종이수납거치대 쑥토리지 15개 벌크 대량 노트북거치 책상수납 오피스 데스크 정리 소품 페이퍼팝 종이가구",
    "cafe24_name": "[대량]아이패드, 노트북 종이수납거치대 쑥토리지 15개입 1박스 벌크 대량할인"
  },
  "11763077838": {
    "cafe24_id": 260,
    "ss_name": "종이수납거치대 쑥토리지 15개 벌크 대량 노트북거치 책상수납 오피스 데스크 정리 소품 페이퍼팝 종이가구",
    "cafe24_name": "[대량]아이패드, 노트북 종이수납거치대 쑥토리지 15개입 1박스 벌크 대량할인"
  },
  "3918260740": {
    "cafe24_id": 86,
    "ss_name": "낮은 미니 옷장 슬림 코너 싱글 간이 원룸 옷장 작은 소형 자취 수납옷장 92cm",
    "cafe24_name": "종이옷장 보야지"
  },
  "9559540220": {
    "cafe24_id": 88,
    "ss_name": "맥북 노트북 거치대 받침대 휴대용 노트북 보관함 종이가구 스탠드",
    "cafe24_name": "노트북 종이수납거치대 쑥토리지"
  },
  "9541795991": {
    "cafe24_id": 85,
    "ss_name": "종이가구 카운터 테이블 인포 안내 데스크 리셉션 계산대 강연대 교탁 강의대 사회대",
    "cafe24_name": "낮은서서 종이책상 테이블"
  },
  "6188143504": {
    "cafe24_id": 127,
    "ss_name": "종이가구 카운터 테이블 인포 안내 데스크 리셉션 계산대 강연대 교탁 강의대 사회대",
    "cafe24_name": "낮은서서 종이책상 테이블"
  },
  "9455409048": {
    "cafe24_id": 201,
    "ss_name": "명함 꽂이 카드 명함 거치대 스탠드 명함 쿠폰 보관함 쿠폰함 3개",
    "cafe24_name": "POP 꽂이 명함 엽서 A4 거치대 (3개입)"
  },
  "5375618240": {
    "cafe24_id": 80,
    "ss_name": "명함 꽂이 카드 명함 거치대 스탠드 명함 쿠폰 보관함 쿠폰함 3개",
    "cafe24_name": "POP 꽂이 명함 엽서 A4 거치대 (3개입)"
  },
  "9455355098": {
    "cafe24_id": 201,
    "ss_name": "엽서 진열대 브로셔 리플렛 거치대 엽서 꽂이 3개",
    "cafe24_name": "POP 꽂이 명함 엽서 A4 거치대 (3개입)"
  },
  "9441002442": {
    "cafe24_id": 201,
    "ss_name": "팜플렛 카달로그 홍보물 거치대 리플렛 꽂이 A4 스탠드 책 도서 진열대 책꽂이 3개",
    "cafe24_name": "POP 꽂이 명함 엽서 A4 거치대 (3개입)"
  },
  "9284554594": {
    "cafe24_id": 203,
    "ss_name": "종이 고양이 화장실 탈부착 속패드",
    "cafe24_name": "고양이 화장실 리필 속박스(3개입)"
  },
  "9031483084": {
    "cafe24_id": 200,
    "ss_name": "페이퍼팝 종이 썬캡 자외선 차단 모자 미소햇 2개입",
    "cafe24_name": "종이모자 미소햇 (2개입)"
  },
  "12242339682": {
    "cafe24_id": 254,
    "ss_name": "페이퍼팝 종이 썬캡 자외선 차단 모자 미소햇 2개입",
    "cafe24_name": "종이모자 미소햇 (2개입)"
  },
  "8927634888": {
    "cafe24_id": 199,
    "ss_name": "분리수거 친환경 휴지통 20개 벌크 대량 도매  병뚜껑수거 20리터 40리터",
    "cafe24_name": "[대량]줍줍박스 20개입 1박스 벌크 대량할인"
  },
  "8828165558": {
    "cafe24_id": 133,
    "ss_name": "종이 스마트폰 거치대 (2개입)",
    "cafe24_name": "종이 스마트폰 거치대 (2개입)"
  },
  "8751031672": {
    "cafe24_id": 197,
    "ss_name": "페이퍼팝 종이 썬캡 여름 햇빛 자외선 차단 썬바이저 모자 대량구매 30개",
    "cafe24_name": "[대량]미소햇 30개 1박스 벌크 대량할인"
  },
  "12990787779": {
    "cafe24_id": 296,
    "ss_name": "페이퍼팝 종이 썬캡 여름 햇빛 자외선 차단 썬바이저 모자 대량구매 30개",
    "cafe24_name": "[대량]미소햇 30개 1박스 벌크 대량할인"
  },
  "8658447402": {
    "cafe24_id": 196,
    "ss_name": "종이가구 이동식트레이 이동트레이 이동수납장 틈새선반 가판대 세탑실수납 간식진열대",
    "cafe24_name": "종이수납장 수수선반 (3단, 4단)"
  },
  "123506348": {
    "cafe24_id": 167,
    "ss_name": "종이가구 이동식트레이 이동트레이 이동수납장 틈새선반 가판대 세탑실수납 간식진열대",
    "cafe24_name": "종이수납장 수수선반 (3단, 4단)"
  },
  "8605101602": {
    "cafe24_id": 195,
    "ss_name": "페이퍼팝 오거나이저 데스크테리어 책상 정리함 수납함 밤부트레이 2개입",
    "cafe24_name": "밤부트레이 (2개입)"
  },
  "648772029": {
    "cafe24_id": 114,
    "ss_name": "페이퍼팝 오거나이저 데스크테리어 책상 정리함 수납함 밤부트레이 2개입",
    "cafe24_name": "밤부트레이 (2개입)"
  },
  "8561651936": {
    "cafe24_id": 83,
    "ss_name": "페이퍼팝 경량 종이 테이블 캠핑 등산 나들이 휴대용",
    "cafe24_name": "휴대용 컵홀더테이블 소소"
  },
  "7287091807": {
    "cafe24_id": 164,
    "ss_name": "페이퍼팝 경량 종이 테이블 캠핑 등산 나들이 휴대용",
    "cafe24_name": "휴대용 컵홀더테이블 소소"
  },
  "8384079012": {
    "cafe24_id": 78,
    "ss_name": "종이 수납장 정리 박스 서랍장 보야지(트렁크49) 도어형",
    "cafe24_name": "종이수납장 보야지트렁크"
  },
  "2010617969": {
    "cafe24_id": 103,
    "ss_name": "종이 수납장 정리 박스 서랍장 보야지(트렁크49) 도어형",
    "cafe24_name": "종이수납장 보야지트렁크"
  },
  "8352027014": {
    "cafe24_id": 183,
    "ss_name": "페이퍼팝 피크닉 페스티벌 캠핑 휴대용 경량 등받이 의자",
    "cafe24_name": "[대량]휴대용 종이의자 테이블 25개 1박스 벌크 대량할인"
  },
  "2634176718": {
    "cafe24_id": 81,
    "ss_name": "페이퍼팝 피크닉 페스티벌 캠핑 휴대용 경량 등받이 의자",
    "cafe24_name": "[대량]휴대용 종이의자 테이블 25개 1박스 벌크 대량할인"
  },
  "8254106011": {
    "cafe24_id": 188,
    "ss_name": "미니 좌식 간이 보조 테이블 거실 사이드 다과 밥상 찻상 페이퍼팝 종이 가구 수수 소반",
    "cafe24_name": "종이테이블 수수소반"
  },
  "12868056931": {
    "cafe24_id": 292,
    "ss_name": "미니 좌식 간이 보조 테이블 거실 사이드 다과 밥상 찻상 페이퍼팝 종이 가구 수수 소반",
    "cafe24_name": "종이테이블 수수소반"
  },
  "8241774420": {
    "cafe24_id": 183,
    "ss_name": "25개 벌크 휴대용 종이의자 판촉 행사 접이식 피크닉 골판지 그라운드체어",
    "cafe24_name": "[대량]휴대용 종이의자 테이블 25개 1박스 벌크 대량할인"
  },
  "8178819726": {
    "cafe24_id": 180,
    "ss_name": "[주문 제작 상담] 허니콤 보드 종이 A박스형  크라프트 양각 가로폭 1000 레터링 전시 행사 팝업 친환경 가벼운 배너 포토존 페이퍼팝",
    "cafe24_name": "[주문제작] 허니콤 종이 레터링"
  },
  "8177587060": {
    "cafe24_id": 179,
    "ss_name": "주문 제작 종이 등신대 전시 행사 팝업 친환경 가벼운 골판지 배너 포토존 입간판 페이퍼팝",
    "cafe24_name": "[주문제작] 등신대"
  },
  "8100446488": {
    "cafe24_id": 176,
    "ss_name": "친환경 재활용 분리수거 종이 휴지통 줍줍박스40L용 추가상품 생분해 봉지 5개입",
    "cafe24_name": "페이퍼팝 생분해 봉지 5개/30개"
  },
  "8056372547": {
    "cafe24_id": 177,
    "ss_name": "친환경 재활용 분리수거 종이 휴지통 줍줍박스40L용 추가상품 생분해 봉지 5개입",
    "cafe24_name": "페이퍼팝 생분해 봉지 5개/30개"
  },
  "8050340502": {
    "cafe24_id": 175,
    "ss_name": "친환경 재활용 분리수거 종이 휴지통 줍줍박스 추가상품 뚜껑",
    "cafe24_name": "줍줍박스 추가상품 뚜껑"
  },
  "7937869548": {
    "cafe24_id": 174,
    "ss_name": "워크샵 화이트보드 이동식 거치대 미니 보드판 가정용 칠판 페이퍼팝 종이가구",
    "cafe24_name": "워크샵 화이트보드"
  },
  "7715430050": {
    "cafe24_id": 169,
    "ss_name": "1인 2인 4인 미니 워크샵 스툴 작업실 거실 보조 낮은 의자 인테리어 종이 페이퍼팝",
    "cafe24_name": "워크샵 스툴"
  },
  "7408078070": {
    "cafe24_id": 166,
    "ss_name": "[라이트책장] 종이책장 종이가구 전면 책꽂이 거실 수납 정리 사무실 장난감",
    "cafe24_name": "[라이트책장] 종이책장 ㅁㅁㅂㅂ"
  },
  "5028145220": {
    "cafe24_id": 84,
    "ss_name": "[라이트책장] 종이책장 종이가구 전면 책꽂이 거실 수납 정리 사무실 장난감",
    "cafe24_name": "[라이트책장] 종이책장 ㅁㅁㅂㅂ"
  },
  "7155894785": {
    "cafe24_id": 163,
    "ss_name": "페이퍼팝 종이소파 다용도 패브릭 소파 스툴 1인 2인 자취 사무실 가구",
    "cafe24_name": "종이소파 보야지"
  },
  "6720573165": {
    "cafe24_id": 82,
    "ss_name": "페스티벌 종이의자 피크닉 나들이 간이의자 조립식 휴대용 토트피크닉세트",
    "cafe24_name": "휴대용 종이 등받이의자 테이블"
  },
  "12569566895": {
    "cafe24_id": 269,
    "ss_name": "페스티벌 종이의자 피크닉 나들이 간이의자 조립식 휴대용 토트피크닉세트",
    "cafe24_name": "휴대용 종이 등받이의자 테이블"
  },
  "6696754458": {
    "cafe24_id": 151,
    "ss_name": "보야지포캣 숨숨집 고양이하우스 트렁크 스크래쳐 종이 페이퍼팝",
    "cafe24_name": "보야지포캣 고양이 숨숨집 스크래쳐"
  },
  "6111333030": {
    "cafe24_id": 88,
    "ss_name": "종이수납거치대 쑥토리지 노트북거치 책상수납 정리 소품 페이퍼팝 종이가구",
    "cafe24_name": "노트북 종이수납거치대 쑥토리지"
  },
  "380598385": {
    "cafe24_id": 89,
    "ss_name": "종이수납거치대 쑥토리지 노트북거치 책상수납 정리 소품 페이퍼팝 종이가구",
    "cafe24_name": "노트북 종이수납거치대 쑥토리지"
  },
  "6090565503": {
    "cafe24_id": 241,
    "ss_name": "실사출력 종이배너  미니 POP 로고출력 스탠드 피켓 주문제작",
    "cafe24_name": "[주문제작] 종이 큐알 스탠드"
  },
  "5910430929": {
    "cafe24_id": 120,
    "ss_name": "종이수납박스 정리함 상자 침대밑 대용량 24/40리터(9개입)",
    "cafe24_name": "수납박스 24L/40L"
  },
  "5910409327": {
    "cafe24_id": 120,
    "ss_name": "종이수납박스 정리함 상자 침대밑 대용량 24/40리터(3개입,6개입)",
    "cafe24_name": "수납박스 24L/40L"
  },
  "5839541820": {
    "cafe24_id": 100,
    "ss_name": "[3단_가로54] 종이책장 페이퍼팝 종이가구 전면 책꽂이 거실 수납 정리 사무실 장난감",
    "cafe24_name": "[3단] 종이책장ㅁㅁㅂㅂ"
  },
  "5839538543": {
    "cafe24_id": 99,
    "ss_name": "전면 책꽂이 종이책장 거실 사무실 장난감 수납 정리 2단 가로54 종이가구",
    "cafe24_name": "[2단] 종이책장ㅁㅁㅂㅂ"
  },
  "5671489171": {
    "cafe24_id": 105,
    "ss_name": "조립식 원룸 간이 싱글 슬림 틈새 코너 미니 주니어 가벼운 옷장 400 페이퍼팝 종이가구",
    "cafe24_name": "종이옷장 보야지"
  },
  "5397202004": {
    "cafe24_id": 85,
    "ss_name": "미니 스탠딩 종이테이블 낮은 미니 책상 행사 진열 가판대",
    "cafe24_name": "낮은서서 종이책상 테이블"
  },
  "5381272536": {
    "cafe24_id": 106,
    "ss_name": "박스제작 가이드북 박스학개론",
    "cafe24_name": "박스제작 가이드북 박스학개론"
  },
  "5277535411": {
    "cafe24_id": 78,
    "ss_name": "종이수납장 보야지 (트렁크98) 티비장 정리 페이퍼팝 종이가구",
    "cafe24_name": "종이수납장 보야지트렁크"
  },
  "5166094474": {
    "cafe24_id": 83,
    "ss_name": "페이퍼팝 야외 피크닉 테이블",
    "cafe24_name": "휴대용 컵홀더테이블 소소"
  },
  "5108489702": {
    "cafe24_id": 238,
    "ss_name": "종이책장 추가 부품 판 추가",
    "cafe24_name": "종이책장 조립 추가부품 세트 (단 연장)"
  },
  "5019801259": {
    "cafe24_id": 95,
    "ss_name": "부품 서서책상 상판",
    "cafe24_name": "부품 서서책상 상판 (추가)"
  },
  "12265142603": {
    "cafe24_id": 309,
    "ss_name": "부품 서서책상 상판",
    "cafe24_name": "부품 서서책상 상판 (추가)"
  },
  "4868628947": {
    "cafe24_id": 80,
    "ss_name": "종이칸막이 맞춤 주문제작",
    "cafe24_name": "[주문제작] 종이칸막이 맞춤"
  },
  "12428193828": {
    "cafe24_id": 78,
    "ss_name": "종이칸막이 맞춤 주문제작",
    "cafe24_name": "[주문제작] 종이칸막이 맞춤"
  },
  "4819051520": {
    "cafe24_id": 97,
    "ss_name": "침대프레임 익스텐션 사이즈 변경 부품",
    "cafe24_name": "침대프레임 익스텐션 사이즈 변경 부품"
  },
  "4755594315": {
    "cafe24_id": 96,
    "ss_name": "부품 캣펀치용 (캣핸드/손잡이)",
    "cafe24_name": "부품 캣펀치용 (캣핸드/손잡이)"
  },
  "4699080337": {
    "cafe24_id": 119,
    "ss_name": "종이 샘플컷팅 및 주문제작",
    "cafe24_name": "종이 샘플컷팅 및 주문제작"
  },
  "4171191301": {
    "cafe24_id": 78,
    "ss_name": "종이수납장 보야지(트렁크49) 도어형 미니 정리박스",
    "cafe24_name": "종이수납장 보야지트렁크"
  },
  "4086186994": {
    "cafe24_id": 104,
    "ss_name": "레터링 스티커 글자컷팅 주문제작 15글자",
    "cafe24_name": "[주문제작] 레터링 스티커 글자컷팅 15글자"
  },
  "12428193829": {
    "cafe24_id": 78,
    "ss_name": "레터링 스티커 글자컷팅 주문제작 15글자",
    "cafe24_name": "[주문제작] 레터링 스티커 글자컷팅 15글자"
  },
  "4077204575": {
    "cafe24_id": 78,
    "ss_name": "종이수납장 보야지(트렁크35) 정리상자 사이드 페이퍼팝 종이가구",
    "cafe24_name": "종이수납장 보야지트렁크"
  },
  "12428193827": {
    "cafe24_id": 78,
    "ss_name": "친환경 재활용 분리수거 종이 휴지통 20리터 40리터 줍줍박스 세트 페이퍼팝 종이가구",
    "cafe24_name": "종이휴지통 줍줍박스(20리터/40리터)"
  },
  "3194832728": {
    "cafe24_id": 238,
    "ss_name": "종이가구 조립 추가부품 PP",
    "cafe24_name": "종이책장 조립 추가부품 세트 (단 연장)"
  },
  "2819178708": {
    "cafe24_id": 82,
    "ss_name": "휴대용 종이의자 메가토트 2개  등받이 접이식 피크닉 세트 골판지 그라운드체어",
    "cafe24_name": "휴대용 종이 등받이의자 테이블"
  },
  "2304349842": {
    "cafe24_id": 93,
    "ss_name": "페이퍼팝 팝업 토이 고양이 장난감 두더지 잡기 펀치 운동 캣토이",
    "cafe24_name": "고양이장난감 캣펀치 (두더지잡기놀이)"
  },
  "424389456": {
    "cafe24_id": 87,
    "ss_name": "파일박스 일부제품 50% 할인",
    "cafe24_name": "파일박스(3개입) 책상위 서류정리 일부제품 50%할인"
  },
  "392791879": {
    "cafe24_id": 92,
    "ss_name": "수납정리함 카인드헬렌(핑크)",
    "cafe24_name": "수납정리함 풋잇백"
  },
  "385417506": {
    "cafe24_id": 198,
    "ss_name": "[20%할인] 공부집중 독서실 칸막이 10개 묶음 (기본/상단/하단)",
    "cafe24_name": "[대량]공부집중 독서실 칸막이 10개 1박스 벌크 대량할인"
  },
  "383376007": {
    "cafe24_id": 89,
    "ss_name": "종이서랍 A (2개입)",
    "cafe24_name": "종이서랍 A/B (2개입) 일부제품 50%할인"
  },
  "324325156": {
    "cafe24_id": 100,
    "ss_name": "종이책장 3단 (가로34) 페이퍼팝 종이가구 전면 책꽂이 거실 수납 정리 사무실 장난감",
    "cafe24_name": "[3단] 종이책장ㅁㅁㅂㅂ"
  },
  "211548656": {
    "cafe24_id": 99,
    "ss_name": "종이책장 전면 책꽂이 2단 가로 34 거실 사무실 장난감 수납 정리 페이퍼팝 종이가구",
    "cafe24_name": "[2단] 종이책장ㅁㅁㅂㅂ"
  },
  "12080117380": {
    "cafe24_id": 264,
    "ss_name": "종이책장 전면 책꽂이 2단 가로 34 거실 사무실 장난감 수납 정리 페이퍼팝 종이가구",
    "cafe24_name": "[2단] 종이책장ㅁㅁㅂㅂ"
  },
  "124097372": {
    "cafe24_id": 115,
    "ss_name": "종이책장 4단 (가로25) 페이퍼팝 종이가구 전면 책꽂이 거실 수납 정리 사무실 장난감",
    "cafe24_name": "[4단] 종이책장ㅁㅁㅂㅂ"
  },
  "124085652": {
    "cafe24_id": 100,
    "ss_name": "종이책장 3단 (가로25) 페이퍼팝 종이가구 전면 책꽂이 거실 수납 정리 사무실 장난감",
    "cafe24_name": "[3단] 종이책장ㅁㅁㅂㅂ"
  },
  "124069601": {
    "cafe24_id": 99,
    "ss_name": "2단 가로44 종이책장 전면 책꽂이 거실 사무실 장난감 수납 정리 페이퍼팝 종이가구",
    "cafe24_name": "[2단] 종이책장ㅁㅁㅂㅂ"
  },
  "12384863196": {
    "cafe24_id": 267,
    "ss_name": "아파트 데스크 서랍 책상 정리함(3개)미니서랍 모듈형 오거나이저 사무실 오피스 수납 소품",
    "cafe24_name": "종이 미니서랍 아파트정리함 (3개입)"
  },
  "13025559151": {
    "cafe24_id": 298,
    "ss_name": "노트북 거치대 스탠드 맥북 휴대용 가벼운 받침대",
    "cafe24_name": "엑스 노트북 받침대"
  },
  "11973620234": {
    "cafe24_id": 263,
    "ss_name": "페이퍼팝 북스탠드 (3개입,10개입) 도서 거치대 서점 도서관 홍보물 책 전시대 북스탠드",
    "cafe24_name": "종이 북스탠드 책거치대(5개입,10개입)"
  },
  "11876091538": {
    "cafe24_id": 261,
    "ss_name": "모니터 받침대 컴퓨터 노트북 듀얼 선반 스탠드",
    "cafe24_name": "플랫 모니터받침대"
  },
  "11762769742": {
    "cafe24_id": 257,
    "ss_name": "고양이 수직 양면 스크래쳐 골판지 반려묘 장난감",
    "cafe24_name": "[회원혜택] 자투리 캣 스크래쳐"
  },
  "12390708256": {
    "cafe24_id": 268,
    "ss_name": "페이퍼팝 위스키 트레이 3개입 홈바 잔 정리 홀더 샘플러 트레이 양주잔 샷잔 칵테일 홀더",
    "cafe24_name": "위스키 트레이 3개입"
  },
  "12491885486": {
    "cafe24_id": 270,
    "ss_name": "라운드 아치형 가림막 파티션 무타공공간분리 접이식 종이 골판지 가벽",
    "cafe24_name": "종이 라운드 파티션"
  },
  "12580051658": {
    "cafe24_id": 273,
    "ss_name": "2026캘린더 도큐멘토X페이퍼팝 탁상 책상용 이젤캘린더",
    "cafe24_name": "2026 도큐멘토 X 페이퍼팝 이젤 캘린더"
  },
  "12585218193": {
    "cafe24_id": 274,
    "ss_name": "고양이 스크래쳐 구멍 놀이 반려묘 장난감 종이 골판지 책꽂이",
    "cafe24_name": "고양이 책꽂이 스크래쳐"
  },
  "12428193821": {
    "cafe24_id": 78,
    "ss_name": "종이수납장 보야지 트렁크 공간박스 다용도 정리함 350x300x400, 1단, 화이트그레이",
    "cafe24_name": "보야지 트렁크"
  },
  "12428193825": {
    "cafe24_id": 78,
    "ss_name": "종이수납장 보야지 트렁크 공간박스 다용도 정리함 490x300x400, 1단, 옐로우",
    "cafe24_name": "보야지 트렁크"
  },
  "12209945860": {
    "cafe24_id": 254,
    "ss_name": "사과상자 공간박스 큐브 수납 보관함 와인 생필품 팬트리 정리함 380x290x230, 1단, 크라프트(3개입)",
    "cafe24_name": "종이수납함 사과상자"
  },
  "11923869459": {
    "cafe24_id": 222,
    "ss_name": "가가77 페이지 X 페이퍼팝 도서진열대",
    "cafe24_name": "페이퍼팝 도서 엽서 진열대"
  },
  "12209949319": {
    "cafe24_id": 254,
    "ss_name": "사과상자 공간박스 큐브 수납 보관함 와인 생필품 팬트리 정리함 380x290x230, 1단, 아이보리(3개입)",
    "cafe24_name": "종이수납함 사과상자"
  }
}

# ── 리뷰 파일 로드 ────────────────────────────────────────────
def load_review(path):
    df = pd.read_excel(path, header=None, dtype=str).fillna('')
    header_idx = 0
    for i, row in df.iterrows():
        if any('상품번호' in str(v) for v in row.values):
            header_idx = i
            break
    headers = list(df.iloc[header_idx])
    data = [list(row) for _, row in df.iloc[header_idx + 1:].iterrows()
            if any(v.strip() for v in row)]
    return headers, data

# ── 변환 로직 ──────────────────────────────────────────────────
OUT_HEADERS = [
    '리뷰_id','상품_id','리뷰_작성_일자','리뷰_작성_시간','리뷰_작성자명',
    '리뷰_제목','리뷰_내용','리뷰_별점','관리자_댓글',
    '구매옵션_옵션명1','구매옵션_옵션값1','구매옵션_옵션명2','구매옵션_옵션값2',
    '구매옵션_옵션명3','구매옵션_옵션값3','구매옵션_옵션명4','구매옵션_옵션값4',
    '구매옵션_옵션명5','구매옵션_옵션값5',
    '고객정보_정보명1','고객정보_답변값1','고객정보_정보명2','고객정보_답변값2',
    '고객정보_정보명3','고객정보_답변값3','고객정보_정보명4','고객정보_답변값4',
    '고객정보_정보명5','고객정보_답변값5',
    'URL_이미지1','URL_이미지2','URL_이미지3','URL_이미지4','URL_이미지5',
    'URL_이미지6','URL_이미지7','URL_이미지8','URL_이미지9','URL_이미지10',
    'URL_동영상1','URL_동영상2','URL_동영상3','URL_동영상4','URL_동영상5',
    'URL_동영상6','URL_동영상7','URL_동영상8','URL_동영상9','URL_동영상10',
]

def convert(mapping, headers, data, date_str):
    def cidx(keyword):
        for i, h in enumerate(headers):
            if keyword in str(h):
                return i
        return -1

    i_ss   = cidx('상품번호')
    i_rate = cidx('평점')
    i_img  = cidx('포토')
    i_cont = cidx('리뷰상세내용')
    i_auth = cidx('등록자')
    i_date = cidx('등록일')

    unmapped = set()
    out_rows = []

    for seq, row in enumerate(data, 1):
        def g(i):
            v = row[i] if 0 <= i < len(row) else None
            return str(v).strip() if v is not None else ''

        ss_num = re.sub(r'\.0$', '', g(i_ss))
        item = mapping.get(ss_num, {})
        cafe24_id = item.get('cafe24_id', '') if item else ''
        if ss_num and not cafe24_id:
            unmapped.add(ss_num)

        raw_date = g(i_date)
        m = re.search(r'(\d{4})\.(\d{2})\.(\d{2})\.\s*(\d{2}:\d{2}:\d{2})', raw_date)
        if m:
            rev_date = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
            rev_time = m.group(4)
        else:
            parts = raw_date.split()
            rev_date = parts[0].replace('.', '-').rstrip('-') if parts else ''
            rev_time = parts[1] if len(parts) > 1 else ''

        img_raw = g(i_img)
        img_urls = [u.strip() for u in re.split(r'[,\n]+', img_raw) if u.strip().startswith('http')]

        out = [''] * len(OUT_HEADERS)
        out[0] = f"{date_str}-{seq:06d}"
        out[1] = str(cafe24_id) if cafe24_id else ''
        out[2] = rev_date
        out[3] = rev_time
        out[4] = g(i_auth)
        out[5] = ''
        out[6] = g(i_cont)
        out[7] = re.sub(r'\.0$', '', g(i_rate))
        for j, url in enumerate(img_urls[:10]):
            out[29 + j] = url

        out_rows.append(out)

    return out_rows, unmapped

def save_xlsx(out_rows, save_path):
    wb = openpyxl.Workbook()
    ws_guide = wb.active
    ws_guide.title = '안내 사항'
    ws_guide['A1'] = '브이리뷰 어드민 > 리뷰 연동 > 리뷰 파일 이관 메뉴에서 업로드해 주세요.'
    ws = wb.create_sheet('review')
    ws.append(OUT_HEADERS)
    for row in out_rows:
        ws.append(row)
    col_widths = {'A': 22, 'B': 12, 'C': 14, 'D': 12, 'E': 16, 'G': 40}
    for col in ws.iter_cols(min_col=30, max_col=49):
        ws.column_dimensions[col[0].column_letter].width = 60
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w
    wb.save(save_path)


# ══════════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════════
PURPLE = '#6c5ce7'
BG     = '#f5f6fa'
CARD   = '#ffffff'

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('브이리뷰 변환기')
        self.resizable(False, False)
        self.configure(bg=BG)

        # 매핑 로드 (저장된 파일 우선, 없으면 내장 데이터)
        saved = load_mapping_from_file()
        self.mapping = saved if saved else dict(BUILTIN_MAPPING)
        if not saved:
            save_mapping_to_file(self.mapping)

        self.review_path    = tk.StringVar()
        self.date_var       = tk.StringVar(value=date.today().strftime('%Y%m%d'))
        self.review_headers = None
        self.review_rows    = None

        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f'+{(sw-w)//2}+{(sh-h)//2}')

    def _build_ui(self):
        header = tk.Frame(self, bg=PURPLE)
        header.pack(fill='x')
        tk.Label(header, text='📋  스마트스토어 → 브이리뷰 변환기',
                 font=('Malgun Gothic', 13, 'bold'), bg=PURPLE, fg='white',
                 pady=14, padx=20).pack(anchor='w')
        tk.Label(header, text='스마트스토어 리뷰 엑셀을 업로드하면 브이리뷰 이관 파일을 만들어드려요',
                 font=('Malgun Gothic', 9), bg=PURPLE, fg='#d9d4ff',
                 pady=0, padx=20).pack(anchor='w')
        tk.Frame(header, height=10, bg=PURPLE).pack()

        nb = ttk.Notebook(self)
        nb.pack(fill='both', expand=True)

        tab1 = tk.Frame(nb, bg=BG)
        tab2 = tk.Frame(nb, bg=BG)
        nb.add(tab1, text='  변환하기  ')
        nb.add(tab2, text='  상품 매핑 관리  ')

        self._build_convert_tab(tab1)
        self._build_mapping_tab(tab2)

    def _build_convert_tab(self, parent):
        outer = tk.Frame(parent, bg=BG)
        outer.pack(fill='both', expand=True, padx=20, pady=16)

        c1 = self._card(outer, '① 스마트스토어 리뷰 파일', '스마트스토어에서 다운받은 리뷰 엑셀')
        row1 = tk.Frame(c1, bg=CARD)
        row1.pack(fill='x', pady=(0, 6))
        self.review_entry = tk.Entry(row1, textvariable=self.review_path,
                                     width=38, relief='flat', bg='#f0eeff',
                                     font=('Malgun Gothic', 9))
        self.review_entry.pack(side='left', ipady=5, padx=(0, 6))
        tk.Button(row1, text='파일 선택', command=self._pick_review,
                  bg=PURPLE, fg='white', relief='flat', cursor='hand2',
                  font=('Malgun Gothic', 9), padx=10).pack(side='left')
        self.review_status = tk.Label(c1, text='', bg=CARD,
                                      font=('Malgun Gothic', 9), fg='#636e72')
        self.review_status.pack(anchor='w')

        c2 = self._card(outer, '② 리뷰 ID 날짜', '리뷰_id 생성에 사용 (예: 20260114-000001)')
        row2 = tk.Frame(c2, bg=CARD)
        row2.pack(fill='x', pady=(0, 2))
        tk.Label(row2, text='날짜 (YYYYMMDD):', bg=CARD,
                 font=('Malgun Gothic', 9), fg='#636e72').pack(side='left', padx=(0, 8))
        tk.Entry(row2, textvariable=self.date_var, width=14,
                 relief='flat', bg='#f0eeff',
                 font=('Malgun Gothic', 10)).pack(side='left', ipady=5)

        self.convert_btn = tk.Button(
            outer, text='⚡  브이리뷰 이관 파일 생성',
            command=self._start_convert,
            bg=PURPLE, fg='white', relief='flat', cursor='hand2',
            font=('Malgun Gothic', 11, 'bold'),
            pady=12, activebackground='#5a4bd1', activeforeground='white'
        )
        self.convert_btn.pack(fill='x', pady=(10, 4))

        self.progress = ttk.Progressbar(outer, mode='indeterminate', length=400)
        self.progress.pack(fill='x', pady=(0, 4))

        self.status_label = tk.Label(outer, text='', bg=BG,
                                     font=('Malgun Gothic', 9), fg='#636e72',
                                     wraplength=440, justify='left')
        self.status_label.pack(anchor='w', pady=(0, 8))

    def _build_mapping_tab(self, parent):
        outer = tk.Frame(parent, bg=BG)
        outer.pack(fill='both', expand=True, padx=20, pady=16)

        top = tk.Frame(outer, bg=BG)
        top.pack(fill='x', pady=(0, 8))
        self.mapping_count_label = tk.Label(top, text=f'현재 등록된 상품: {len(self.mapping)}개',
                                             bg=BG, font=('Malgun Gothic', 10, 'bold'), fg=PURPLE)
        self.mapping_count_label.pack(side='left')

        search_frame = tk.Frame(outer, bg=BG)
        search_frame.pack(fill='x', pady=(0, 8))
        tk.Label(search_frame, text='검색:', bg=BG, font=('Malgun Gothic', 9)).pack(side='left', padx=(0, 6))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *a: self._refresh_table())
        tk.Entry(search_frame, textvariable=self.search_var, width=25,
                 relief='flat', bg='#f0eeff', font=('Malgun Gothic', 9)).pack(side='left', ipady=4)

        table_frame = tk.Frame(outer, bg=BG)
        table_frame.pack(fill='both', expand=True)

        cols = ('ss_num', 'cafe24_id', 'cafe24_name')
        self.tree = ttk.Treeview(table_frame, columns=cols, show='headings', height=12)
        self.tree.heading('ss_num',      text='스스 상품번호')
        self.tree.heading('cafe24_id',   text='카페24 ID')
        self.tree.heading('cafe24_name', text='카페24 상품명')
        self.tree.column('ss_num',      width=140, anchor='center')
        self.tree.column('cafe24_id',   width=90,  anchor='center')
        self.tree.column('cafe24_name', width=220)

        sb = ttk.Scrollbar(table_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side='left', fill='both', expand=True)
        sb.pack(side='left', fill='y')
        self._refresh_table()

        btn_frame = tk.Frame(outer, bg=BG)
        btn_frame.pack(side='left', fill='y', padx=(12, 0))

        def btn(text, cmd, color=PURPLE):
            tk.Button(btn_frame, text=text, command=cmd,
                      bg=color, fg='white', relief='flat', cursor='hand2',
                      font=('Malgun Gothic', 9), padx=8, pady=6,
                      width=10).pack(fill='x', pady=3)

        btn('➕ 추가', self._add_item)
        btn('✏️ 수정', self._edit_item)
        btn('🗑 삭제', self._delete_item, color='#e17055')
        tk.Frame(btn_frame, bg=BG, height=20).pack()
        btn('💾 저장', self._save_mapping, color='#00b894')

    def _refresh_table(self):
        keyword = self.search_var.get().lower() if hasattr(self, 'search_var') else ''
        for row in self.tree.get_children():
            self.tree.delete(row)
        for ss_num, item in sorted(self.mapping.items(), key=lambda x: int(x[1].get('cafe24_id', 0))):
            c24_id   = item.get('cafe24_id', '')
            c24_name = item.get('cafe24_name', '')
            if keyword and keyword not in ss_num and keyword not in str(c24_id) and keyword not in c24_name.lower():
                continue
            self.tree.insert('', 'end', values=(ss_num, c24_id, c24_name))
        self.mapping_count_label.config(text=f'현재 등록된 상품: {len(self.mapping)}개')

    def _add_item(self):
        self._open_edit_dialog()

    def _edit_item(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning('알림', '수정할 항목을 선택해주세요.')
            return
        vals = self.tree.item(sel[0])['values']
        self._open_edit_dialog(ss_num=str(vals[0]), cafe24_id=str(vals[1]), cafe24_name=str(vals[2]))

    def _delete_item(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning('알림', '삭제할 항목을 선택해주세요.')
            return
        vals = self.tree.item(sel[0])['values']
        ss_num = str(vals[0])
        if messagebox.askyesno('확인', f'스스 번호 {ss_num} 을 삭제할까요?'):
            self.mapping.pop(ss_num, None)
            self._refresh_table()

    def _save_mapping(self):
        save_mapping_to_file(self.mapping)
        messagebox.showinfo('저장 완료', f'{len(self.mapping)}개 상품 매핑이 저장되었습니다!\n\n저장 위치: {get_mapping_path()}')

    def _open_edit_dialog(self, ss_num='', cafe24_id='', cafe24_name=''):
        dlg = tk.Toplevel(self)
        dlg.title('상품 추가/수정')
        dlg.configure(bg=CARD)
        dlg.resizable(False, False)
        dlg.grab_set()

        x = self.winfo_x() + (self.winfo_width() - 360) // 2
        y = self.winfo_y() + (self.winfo_height() - 200) // 2
        dlg.geometry(f'360x210+{x}+{y}')

        fields = [
            ('스마트스토어 상품번호', ss_num),
            ('카페24 상품 ID',        cafe24_id),
            ('카페24 상품명',          cafe24_name),
        ]
        entries = []
        for label, val in fields:
            row = tk.Frame(dlg, bg=CARD)
            row.pack(fill='x', padx=20, pady=6)
            tk.Label(row, text=label, bg=CARD, width=18, anchor='w',
                     font=('Malgun Gothic', 9)).pack(side='left')
            e = tk.Entry(row, relief='flat', bg='#f0eeff',
                         font=('Malgun Gothic', 10), width=20)
            e.insert(0, val)
            e.pack(side='left', ipady=4)
            entries.append(e)

        is_edit = bool(ss_num)

        def save():
            new_ss   = entries[0].get().strip()
            new_c24  = entries[1].get().strip()
            new_name = entries[2].get().strip()
            if not new_ss or not new_c24:
                messagebox.showwarning('알림', '상품번호와 카페24 ID는 필수예요.', parent=dlg)
                return
            try:
                int(new_c24)
            except ValueError:
                messagebox.showwarning('알림', '카페24 ID는 숫자여야 해요.', parent=dlg)
                return
            if is_edit and ss_num != new_ss:
                self.mapping.pop(ss_num, None)
            self.mapping[new_ss] = {'cafe24_id': int(new_c24), 'cafe24_name': new_name, 'ss_name': ''}
            self._refresh_table()
            dlg.destroy()

        tk.Button(dlg, text='저장', command=save,
                  bg=PURPLE, fg='white', relief='flat', cursor='hand2',
                  font=('Malgun Gothic', 10, 'bold'), pady=8).pack(fill='x', padx=20, pady=10)

    def _pick_review(self):
        path = filedialog.askopenfilename(
            title='스마트스토어 리뷰 파일 선택',
            filetypes=[('Excel', '*.xlsx *.xls'), ('All', '*.*')]
        )
        if not path:
            return
        self.review_path.set(path)
        self.review_status.config(text='⏳ 읽는 중...', fg='#a29bfe')
        self.update()
        try:
            self.review_headers, self.review_rows = load_review(path)
            self.review_status.config(text=f'✅ {len(self.review_rows)}건 로드 완료', fg='#00b894')
        except Exception as e:
            self.review_headers = self.review_rows = None
            self.review_status.config(text=f'❌ 오류: {e}', fg='#d63031')

    def _start_convert(self):
        if not self.review_rows:
            messagebox.showwarning('알림', '리뷰 파일을 먼저 선택해주세요.')
            return
        date_str = self.date_var.get().strip().replace('-', '')
        if len(date_str) != 8 or not date_str.isdigit():
            messagebox.showwarning('알림', '날짜를 YYYYMMDD 형식으로 입력해주세요.')
            return
        save_path = filedialog.asksaveasfilename(
            title='저장 위치 선택',
            defaultextension='.xlsx',
            initialfile=f'브이리뷰_이관파일_{date_str}.xlsx',
            filetypes=[('Excel', '*.xlsx')]
        )
        if not save_path:
            return

        self.convert_btn.config(state='disabled')
        self.progress.start(10)
        self.status_label.config(text='변환 중...', fg='#a29bfe')

        def run():
            try:
                out_rows, unmapped = convert(
                    self.mapping, self.review_headers,
                    self.review_rows, date_str
                )
                save_xlsx(out_rows, save_path)
                self.after(0, lambda: self._done(len(out_rows), unmapped, save_path))
            except Exception as e:
                self.after(0, lambda: self._error(str(e)))

        threading.Thread(target=run, daemon=True).start()

    def _done(self, count, unmapped, save_path):
        self.progress.stop()
        self.convert_btn.config(state='normal')
        msg = f'✅ {count}건 변환 완료 → {os.path.basename(save_path)}'
        if unmapped:
            msg += f'\n⚠️ 매핑 불가 상품번호 {len(unmapped)}개: {", ".join(list(unmapped)[:5])}'
            if len(unmapped) > 5:
                msg += ' ...'
        self.status_label.config(text=msg, fg='#00b894')
        messagebox.showinfo('완료', f'{count}건 변환 완료!\n저장 위치: {save_path}')

    def _error(self, msg):
        self.progress.stop()
        self.convert_btn.config(state='normal')
        self.status_label.config(text=f'❌ 오류: {msg}', fg='#d63031')
        messagebox.showerror('오류', msg)

    def _card(self, parent, title, subtitle=''):
        frame = tk.Frame(parent, bg=CARD, bd=0,
                         highlightthickness=1, highlightbackground='#e9e4ff')
        frame.pack(fill='x', pady=(0, 12))
        inner = tk.Frame(frame, bg=CARD)
        inner.pack(fill='x', padx=16, pady=12)
        tk.Label(inner, text=title, bg=CARD,
                 font=('Malgun Gothic', 10, 'bold'), fg=PURPLE).pack(anchor='w')
        if subtitle:
            tk.Label(inner, text=subtitle, bg=CARD,
                     font=('Malgun Gothic', 8), fg='#b2bec3').pack(anchor='w', pady=(1, 8))
        return inner


if __name__ == '__main__':
    app = App()
    app.mainloop()
